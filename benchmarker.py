# -*- coding: utf-8 -*-
#################### IMPORTING DEPENDENCIES ####################
################### INSTALLING DEPENDENCIES ############
import subprocess
import sys

# List of required libraries
required_libraries = [
    "numpy", "scikit-learn", "tensorflow", "pandas", "openpyxl"
]

# Function to install missing libraries
def install_libraries():
    for library in required_libraries:
        try:
            __import__(library)
        except ImportError:
            print(f"{library} not found. Installing...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", library])
        else:
            print(f"{library} is already installed.")

# Run the installation check
install_libraries()


import numpy as np
from sklearn.model_selection import train_test_split, cross_val_score
from tensorflow.keras.datasets import mnist, fashion_mnist
import pandas as pd
import gzip
import pickle
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import tensorflow as tf
import random
import time
from sklearn.metrics import accuracy_score
from sklearn.base import clone
from sklearn.preprocessing import StandardScaler
from sklearn.tree import DecisionTreeClassifier # for testing
from tensorflow.keras.models import load_model


#################### DEFINING FUNCTIONS ####################
# Function to sample a subset of data
def sample_data(data, labels, num_samples=10000):
    indices = np.random.choice(len(data), num_samples, replace=False)
    return data[indices], labels[indices]
# Benchmarking function
def benchmark_model(model, X, y, model_name, dataset_name, repeats=3, is_tf_model=False):
    results = []
    for _ in range(repeats):
        # Split the data
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        # Preprocess for TensorFlow models
        if is_tf_model:
            X_train = X_train.reshape(-1, 28, 28, 1).astype("float32") / 255.0
            X_test = X_test.reshape(-1, 28, 28, 1).astype("float32") / 255.0

        # Skip cloning for TensorFlow/Keras models
        if not is_tf_model:
            from sklearn.base import clone
            model = clone(model)

        # Measure training time
        start_time = time.time()
        if is_tf_model:  # Fit and predict for TensorFlow/Keras models
            model.fit(X_train, y_train, epochs=1, batch_size=32, verbose=0)
            y_pred = model.predict(X_test).argmax(axis=1)
        else:  # Fit and predict for Scikit-learn models
            model.fit(X_train, y_train)
            y_pred = model.predict(X_test)
        train_time = time.time() - start_time

        # Calculate accuracy
        accuracy = accuracy_score(y_test, y_pred)
        results.append((accuracy, train_time))

    # Calculate mean and std
    accuracies = [r[0] for r in results]
    times = [r[1] for r in results]

    return {
        "Model": model_name,
        "Dataset": dataset_name,
        "Accuracy (mean)": np.mean(accuracies),
        "Accuracy (std)": np.std(accuracies),
        "Training Time (mean)": np.mean(times),
    }

def run_benchmark(models, datasets, repeats=3):
    benchmark_results = []
    for model_name, model in models.items():
        is_tf_model = isinstance(model, tf.keras.Model)  # Check if it's a Keras model
        for dataset_name, (X, y) in datasets.items():
            print(f"Running benchmark for {model_name} on {dataset_name}...")
            result = benchmark_model(
                model, X, y, model_name, dataset_name, repeats, is_tf_model
            )
            benchmark_results.append(result)

    # Convert results to a DataFrame
    results_df = pd.DataFrame(benchmark_results)
    return results_df

def scale_dataset(X, y):
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    return X_scaled, y

# Modified function to format results with merged headers
def format_results_table_with_merged_headers(results_df):
    # Create hierarchical MultiIndex for columns
    columns = pd.MultiIndex.from_tuples([
        ("Model", ""),
        ("MNIST", "accuracy (mean)"),
        ("MNIST", "accuracy (std)"),
        ("MNIST", "Training Time (mean)"),
        ("Fashion-MNIST", "accuracy (mean)"),
        ("Fashion-MNIST", "accuracy (std)"),
        ("Fashion-MNIST", "Training Time (mean)"),
        ("Animal Silhouette MNIST", "accuracy (mean)"),
        ("Animal Silhouette MNIST", "accuracy (std)"),
        ("Animal Silhouette MNIST", "Training Time (mean)")
    ])

    # Initialize an empty DataFrame with hierarchical columns
    formatted_results = pd.DataFrame(columns=columns)

    # Fill the rows
    for model in results_df["Model"].unique():
        row = [model]  # Start with the model name
        for dataset in ["MNIST", "Fashion-MNIST", "Custom Dataset"]:
            dataset_name = "Animal Silhouette MNIST" if dataset == "Custom Dataset" else dataset
            subset = results_df[(results_df["Model"] == model) & (results_df["Dataset"] == dataset)]
            if not subset.empty:
                row.extend([subset["Accuracy (mean)"].values[0],
                            subset["Accuracy (std)"].values[0],
                            subset["Training Time (mean)"].values[0]])
            else:
                row.extend([None, None, None])  # Fill empty values if no results
        formatted_results.loc[len(formatted_results)] = row

    return formatted_results


def load_animal_shape(data_path, label_path):
    with gzip.GzipFile(data_path, 'rb') as f:
        data = pickle.load(f)
    with gzip.GzipFile(label_path, 'rb') as f:
        label = pickle.load(f)
    return data, label

def load_uploaded_model(file_path):
    # Check file extension
    if file_path.endswith(".h5") or file_path.endswith(".keras"):
        print("Loading TensorFlow/Keras model...")
        model = load_model(file_path)
    elif file_path.endswith(".pkl"):
        print("Loading pickle model...")
        with open(file_path, "rb") as f:
            model = pickle.load(f)
    else:
        raise ValueError("Unsupported file format. Please upload a .keras, .h5, or .pkl file.")
    return model

def create_table():
  # Create a workbook and add a worksheet
  wb = Workbook()
  ws = wb.active

  # Write the merged headers
  headers = [
      ("Model", ""),
      ("MNIST", "accuracy (mean)"),
      ("MNIST", "accuracy (std)"),
      ("MNIST", "Training Time (mean)"),
      ("Fashion-MNIST", "accuracy (mean)"),
      ("Fashion-MNIST", "accuracy (std)"),
      ("Fashion-MNIST", "Training Time (mean)"),
      ("Animal Silhouette MNIST", "accuracy (mean)"),
      ("Animal Silhouette MNIST", "accuracy (std)"),
      ("Animal Silhouette MNIST", "Training Time (mean)")
  ]

  # Merge cells and write dataset headers
  col = 2
  for dataset in ["MNIST", "Fashion-MNIST", "Animal Silhouette MNIST"]:
      ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
      ws.cell(row=1, column=col).value = dataset
      ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
      ws.cell(row=1, column=col).font = Font(bold=True)
      col += 3

  # Write subheaders
  for i, header in enumerate(headers):
      cell = ws.cell(row=2, column=i + 1)
      cell.value = header[1] if header[1] else header[0]
      cell.alignment = Alignment(horizontal="center", vertical="center")
      cell.font = Font(bold=True)

  # Write the data
  for i, row in enumerate(formatted_results.itertuples(index=False), start=3):
      for j, value in enumerate(row, start=1):
          cell = ws.cell(row=i, column=j)
          cell.value = value
          cell.alignment = Alignment(horizontal="center", vertical="center")

  # Add borders to the table
  thin_border = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )

  for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
      for cell in row:
          cell.border = thin_border

  # Adjust column widths
  for col_idx in range(1, ws.max_column + 1):
      col_letter = get_column_letter(col_idx)
      ws.column_dimensions[col_letter].width = 20  # Set column width

  # Save the styled table
  output_folder = os.path.join(os.getcwd(), "output")
  # Create the folder if it doesn't exist
  if not os.path.exists(output_folder):
      os.makedirs(output_folder)
  # Define the full path for the output Excel file
  output_file_path = os.path.join(output_folder, "Benchmark_file.xlsx")
  wb.save(output_file_path)
  print(f" table saved to {output_file_path}")


#################### LOADING THE DATASETS ####################
# Load MNIST
(x_train_mnist, y_train_mnist), (x_test_mnist, y_test_mnist) = mnist.load_data()
x_mnist, y_mnist = sample_data(np.concatenate([x_train_mnist, x_test_mnist]),
                               np.concatenate([y_train_mnist, y_test_mnist]),
                               num_samples=10000)
# Load Fashion-MNIST
(x_train_fashion, y_train_fashion), (x_test_fashion, y_test_fashion) = fashion_mnist.load_data()
x_fashion, y_fashion = sample_data(np.concatenate([x_train_fashion, x_test_fashion]),
                                   np.concatenate([y_train_fashion, y_test_fashion]),
                                   num_samples=10000)

expected_files = [
    os.path.join("data", "Animal Shape", "animal_data_version_3.gz"),
    os.path.join("data", "Animal Shape", "animal_label_version_3.gz")
]

file_paths = {}
for expected_file in expected_files:
    while True:
        if os.path.exists(expected_file):
            print(f"Data file found at: {expected_file}\n\n")
            file_paths[os.path.basename(expected_file)] = expected_file
            break
        else:
            print(f"{os.path.basename(expected_file)} not found!!\n")
            file_path = input(f'Please input the location of "{os.path.basename(expected_file)}": ')
            if os.path.isfile(file_path) and os.path.basename(file_path) == os.path.basename(expected_file):
                print(f"Data file found at: {file_path}\n\n")
                file_paths[os.path.basename(expected_file)] = file_path
                break
            else:
                print(f"Error: The file at {file_path} does not match '{os.path.basename(expected_file)}'. Please check the path and try again.\n")

# If necassary, the location can be changed manually
DATA_PATH = file_paths["animal_data_version_3.gz"]
LABEL_PATH = file_paths["animal_label_version_3.gz"]

# Loading the Animal-MNIST dataset
X_custom, y_custom = load_animal_shape(DATA_PATH, LABEL_PATH)

# Display information
print(f"MNIST data shape: {x_mnist.shape}, labels shape: {y_mnist.shape}")
print(f"Fashion-MNIST data shape: {x_fashion.shape}, labels shape: {y_fashion.shape}")
print(f"Animal dataset shape: {X_custom.shape}, labels shape: {y_custom.shape}")

CHECK = input("Would you like the datasets to be scaled?: ")
if CHECK.lower() == "yes" or CHECK.lower() =="y":
  # Scale datasets

  datasets = {
      "MNIST": scale_dataset(x_mnist.reshape(x_mnist.shape[0], -1), y_mnist),
      "Fashion-MNIST": scale_dataset(x_fashion.reshape(x_fashion.shape[0], -1), y_fashion),
      "Custom Dataset": scale_dataset(X_custom.reshape(X_custom.shape[0], -1), y_custom)
  }
else:
  datasets = {
    "MNIST": (x_mnist, y_mnist),
    "Fashion-MNIST": (x_fashion, y_fashion),
    "Custom Dataset": (X_custom, y_custom)
  }
#################### LOADING THE USER MODEL ####################
MODEL_NAME = input("Please enter the name of your Model: ")
MODEL_PATH = input("Please enter the location to your Model: ")
CHECK = input("Would you like to see the comparison between your Model and Random Tree(\"entropy\", max_depth=10, \"best\")?: ")
user_model = load_uploaded_model (MODEL_PATH)

if CHECK.lower() == "yes" or CHECK.lower() =="y":
  models = {
      MODEL_NAME: user_model,
      "Random Tree": DecisionTreeClassifier(
          criterion="entropy", max_depth=10, splitter="best", random_state=42
      )
  }
else:
      models = {
      MODEL_NAME: user_model
  }

#################### RUNNING BENCHMARK ####################
results_df = run_benchmark(models, datasets)
formatted_results = format_results_table_with_merged_headers(results_df)
create_table()
